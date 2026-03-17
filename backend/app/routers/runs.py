from __future__ import annotations
import io
import json
import uuid
from typing import List, Optional

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from ..config import settings
from ..database import get_db
from ..models.audience import Persona
from ..models.experiment import Experiment, ExperimentDistVariable, ExperimentVariable, OutputSchema, Question, SynonymSet
from ..models.simulation import SimulationRun, SimulationTask
from ..schemas.simulation import (
    LaunchRequest, ReExtractRequest,
    SimulationRunOut, SimulationTaskDetail, SimulationTaskSummary,
)
from ..services.execution import launch_run, re_extract_run, retry_failed_tasks
from ..services.sampling import sample_correlated_population
from ..services.backstory import generate_backstory
from ..services.variable_resolution import resolve_variables, resolve_dist_variables, apply_synonym_injection
from ..services.validation import validate_persona

router = APIRouter(prefix="/api", tags=["runs"])


# ── List runs ────────────────────────────────────────────────────────────────

@router.get("/runs", response_model=List[SimulationRunOut])
async def list_runs(
    experiment_id: Optional[str] = Query(None),
    limit: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
):
    """List simulation runs, optionally filtered by experiment."""
    q = select(SimulationRun).order_by(SimulationRun.created_at.desc()).limit(limit)
    if experiment_id:
        q = q.where(SimulationRun.experiment_id == experiment_id)
    result = await db.execute(q)
    return result.scalars().all()


# ── Launch ────────────────────────────────────────────────────────────────────

@router.post("/experiments/{experiment_id}/launch", response_model=SimulationRunOut, status_code=202)
async def launch_simulation(
    experiment_id: str,
    body: LaunchRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
):
    """
    Launch a simulation run. Returns 202 immediately; execution runs in background.
    Requires the X-Confirm: true header to prevent accidental launches.
    """
    exp = await db.get(Experiment, experiment_id)
    if not exp:
        raise HTTPException(status_code=404, detail="Experiment not found")

    # ── Resolve which personas to use ────────────────────────────────────────
    import random

    async def _make_fresh_personas(n: int) -> list[str]:
        """Sample n new personas WITHOUT backstory (fast) and persist them.
        Backstories are generated lazily inside launch_run before task execution."""
        raw_traits = await sample_correlated_population(exp.audience_id, n, db)
        ids: list[str] = []
        for traits in raw_traits:
            score, _ = validate_persona(traits)
            p = Persona(
                id=str(uuid.uuid4()),
                audience_id=exp.audience_id,
                traits_json=traits,
                backstory=None,   # generated in background before first use
                plausibility=score,
                flagged=score < settings.plausibility_threshold,
            )
            db.add(p)
            ids.append(p.id)
        return ids

    if body.persona_ids:
        # Explicit list supplied by caller
        persona_ids = body.persona_ids[: body.population_size]
        if len(persona_ids) < body.population_size:
            raise HTTPException(status_code=400, detail="Not enough persona_ids for the requested population_size")

    elif body.sample_fresh:
        # User explicitly asked for a new sample
        persona_ids = await _make_fresh_personas(body.population_size)

    else:
        # Default: reuse existing personas already sampled for this audience
        existing_result = await db.execute(
            select(Persona)
            .where(Persona.audience_id == exp.audience_id)
            .where(Persona.flagged == False)  # noqa: E712
        )
        existing = existing_result.scalars().all()

        if existing:
            chosen = random.sample(existing, min(body.population_size, len(existing)))
            persona_ids = [p.id for p in chosen]
            shortage = body.population_size - len(persona_ids)
            if shortage > 0:
                # Not enough personas in pool — top up with fresh ones
                persona_ids += await _make_fresh_personas(shortage)
        else:
            # Audience has no personas yet — sample fresh automatically
            persona_ids = await _make_fresh_personas(body.population_size)

    # Resolve variables for each persona to get injected_vars
    vars_result = await db.execute(
        select(ExperimentVariable).where(ExperimentVariable.experiment_id == experiment_id)
    )
    exp_vars = vars_result.scalars().all()

    dist_vars_result = await db.execute(
        select(ExperimentDistVariable).where(ExperimentDistVariable.experiment_id == experiment_id)
    )
    dist_vars_for_task = dist_vars_result.scalars().all()

    q_result = await db.execute(
        select(Question).where(Question.experiment_id == experiment_id)
    )
    questions_for_vars = q_result.scalars().all()

    syn_result = await db.execute(
        select(SynonymSet).where(SynonymSet.experiment_id == experiment_id)
    )
    synonym_sets = syn_result.scalars().all()

    # Load output schema for locked_config
    schema_result = await db.execute(
        select(OutputSchema)
        .where(OutputSchema.experiment_id == experiment_id)
        .order_by(OutputSchema.version.desc())
    )
    schema = schema_result.scalars().first()
    output_schema = schema.schema_json if schema else []

    # Create run
    locked_config = {
        "experiment_id": experiment_id,
        "audience_id": exp.audience_id,
        "execution_mode": exp.execution_mode,
        "dual_extraction": body.dual_extraction,
        "output_schema": output_schema,
        "synonym_injection_enabled": exp.synonym_injection_enabled,
        "drift_detection_enabled": exp.drift_detection_enabled,
        "est_tokens_per_task": 2000,
    }

    run = SimulationRun(
        id=str(uuid.uuid4()),
        experiment_id=experiment_id,
        model_pass1=body.model_pass1 or settings.model_pass1,
        model_pass2=body.model_pass2 or settings.model_pass2,
        status="pending",
        total_tasks=body.population_size,
        locked_config=locked_config,
    )
    db.add(run)

    # Create tasks — resolve ALL variable placeholders (global_context + every question
    # text) into a single cache so injected_vars matches exactly what the LLM receives.
    for pid in persona_ids:
        resolved_cache: dict[str, str] = {}
        dist_cache: dict[str, str] = {}
        # Scan global context
        resolve_variables(exp.global_context, exp_vars, resolved_cache)
        resolve_dist_variables(exp.global_context, dist_vars_for_task, dist_cache)
        # Scan every question text so var_* captures all placeholders, not just those in context
        for q in questions_for_vars:
            resolve_variables(q.question_text, exp_vars, resolved_cache)
            resolve_dist_variables(q.question_text, dist_vars_for_task, dist_cache)
        task = SimulationTask(
            id=str(uuid.uuid4()),
            run_id=run.id,
            persona_id=pid,
            injected_vars={**resolved_cache, **dist_cache},
        )
        db.add(task)

    await db.commit()
    await db.refresh(run)

    # Launch background
    background_tasks.add_task(launch_run, run.id, None)

    return run


# ── Run status & tasks ────────────────────────────────────────────────────────

@router.get("/runs/{run_id}", response_model=SimulationRunOut)
async def get_run(run_id: str, db: AsyncSession = Depends(get_db)):
    run = await db.get(SimulationRun, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    return run


@router.get("/runs/{run_id}/progress")
async def get_run_progress(run_id: str, db: AsyncSession = Depends(get_db)):
    """Lightweight task-level progress counts — used by the sidebar tracker."""
    from sqlalchemy import func
    result = await db.execute(
        select(
            func.count().filter(SimulationTask.pass1_status == "running").label("p1_running"),
            func.count().filter(SimulationTask.pass1_status == "completed").label("p1_done"),
            func.count().filter(SimulationTask.pass2_status == "running").label("p2_running"),
            func.count().filter(SimulationTask.pass2_status == "completed").label("p2_done"),
            func.count().filter(SimulationTask.pass1_status == "failed").label("failed"),
            func.count().label("total"),
        ).where(SimulationTask.run_id == run_id)
    )
    row = result.one()
    return {
        "p1_running": row.p1_running,
        "p1_done":    row.p1_done,
        "p2_running": row.p2_running,
        "p2_done":    row.p2_done,
        "failed":     row.failed,
        "total":      row.total,
    }


@router.get("/runs/{run_id}/tasks", response_model=List[SimulationTaskSummary])
async def list_tasks(
    run_id: str,
    offset: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(SimulationTask)
        .where(SimulationTask.run_id == run_id)
        .offset(offset)
        .limit(limit)
    )
    return result.scalars().all()


@router.get("/runs/{run_id}/tasks/{task_id}", response_model=SimulationTaskDetail)
async def get_task(run_id: str, task_id: str, db: AsyncSession = Depends(get_db)):
    task = await db.get(SimulationTask, task_id)
    if not task or task.run_id != run_id:
        raise HTTPException(status_code=404, detail="Task not found")
    return task


# ── Actions ───────────────────────────────────────────────────────────────────

@router.post("/runs/{run_id}/retry-failed", status_code=202)
async def retry_run(run_id: str, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_db)):
    run = await db.get(SimulationRun, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    background_tasks.add_task(retry_failed_tasks, run_id)
    return {"status": "retry_queued", "run_id": run_id}


@router.post("/runs/{run_id}/cancel", status_code=200)
async def cancel_run(run_id: str, db: AsyncSession = Depends(get_db)):
    run = await db.get(SimulationRun, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    run.status = "cancelled"
    await db.commit()
    return {"status": "cancelled", "run_id": run_id}


@router.delete("/runs", status_code=200)
async def prune_runs(
    experiment_id: Optional[str] = Query(None),
    include_completed: bool = Query(False),
    db: AsyncSession = Depends(get_db),
):
    """
    Delete finished runs (optionally scoped to one experiment).
    - include_completed=false (default): deletes failed + cancelled only
    - include_completed=true:            deletes failed + cancelled + completed
    Running and pending runs are never touched.
    """
    statuses = ["failed", "cancelled"]
    if include_completed:
        statuses.append("completed")
    q = select(SimulationRun).where(SimulationRun.status.in_(statuses))
    if experiment_id:
        q = q.where(SimulationRun.experiment_id == experiment_id)
    result = await db.execute(q)
    runs = result.scalars().all()
    run_ids = [r.id for r in runs]
    if run_ids:
        await db.execute(delete(SimulationTask).where(SimulationTask.run_id.in_(run_ids)))
        for run in runs:
            await db.delete(run)
        await db.commit()
    return {"deleted": len(run_ids)}


@router.delete("/runs/{run_id}", status_code=204)
async def delete_run(run_id: str, db: AsyncSession = Depends(get_db)):
    """
    Delete a single run and all its tasks.
    Only allowed when the run is in 'completed' or 'cancelled' status.
    """
    run = await db.get(SimulationRun, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    if run.status not in ("completed", "cancelled", "failed"):
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete run with status '{run.status}'. "
                   "Only completed, cancelled, or failed runs may be deleted.",
        )
    await db.execute(delete(SimulationTask).where(SimulationTask.run_id == run_id))
    await db.delete(run)
    await db.commit()


@router.post("/runs/{run_id}/resume", status_code=202)
async def resume_run(run_id: str, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_db)):
    """Resume a cancelled run — re-runs all pending tasks."""
    run = await db.get(SimulationRun, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    if run.status not in ("cancelled", "failed"):
        raise HTTPException(status_code=400, detail=f"Run cannot be resumed (status: {run.status})")
    run.status = "running"
    await db.commit()
    background_tasks.add_task(launch_run, run_id, None)
    return {"status": "resumed", "run_id": run_id}


@router.post("/runs/{run_id}/re-extract", status_code=202)
async def re_extract(
    run_id: str,
    body: ReExtractRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
):
    run = await db.get(SimulationRun, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    background_tasks.add_task(re_extract_run, run_id, body.schema_version)
    return {"status": "re_extract_queued", "run_id": run_id}


# ── Export ────────────────────────────────────────────────────────────────────

@router.get("/runs/{run_id}/export")
async def export_run(
    run_id: str,
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    include_confidence: bool = Query(True),
    include_transcript: bool = Query(False),
    db: AsyncSession = Depends(get_db),
):
    run = await db.get(SimulationRun, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")

    tasks_result = await db.execute(
        select(SimulationTask).where(SimulationTask.run_id == run_id)
    )
    tasks = tasks_result.scalars().all()

    # Get calibration level for badge
    from ..models.simulation import CalibrationStatus
    cal = await db.get(CalibrationStatus, run.experiment_id)
    cal_level = cal.level if cal else "uncalibrated"

    rows = []
    for task in tasks:
        row: dict = {
            "task_id": task.id,
            "persona_id": task.persona_id,
            "pass1_status": task.pass1_status,
            "pass2_status": task.pass2_status,
            "drift_flagged": task.drift_flagged,
            "pass1_cost_usd": task.pass1_cost_usd,
            "pass2_cost_usd": task.pass2_cost_usd,
            "calibration_badge": cal_level,
        }
        # Injected variables — the {{placeholder}} values resolved for this persona/task.
        # Prefixed with "var_" so they are visually grouped and don't clash with output fields.
        if task.injected_vars:
            for k, v in task.injected_vars.items():
                row[f"var_{k}"] = v
        if task.extracted_json:
            row.update(task.extracted_json)
        if include_confidence and task.extraction_confidence:
            row.update({f"{k}_confidence": v for k, v in task.extraction_confidence.items()})
        if include_transcript:
            row["transcript"] = task.raw_transcript
        rows.append(row)

    if format == "csv":
        if not rows:
            return StreamingResponse(io.StringIO(""), media_type="text/csv")
        headers = list(rows[0].keys())
        buf = io.StringIO()
        buf.write(",".join(headers) + "\n")
        for row in rows:
            buf.write(",".join(str(row.get(h, "")) for h in headers) + "\n")
        buf.seek(0)
        return StreamingResponse(
            io.BytesIO(buf.getvalue().encode()),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=run_{run_id}.csv"},
        )
    else:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        # Always include pass1_prompt and transcript in XLSX
        task_map = {t.id: t for t in tasks}
        for row in rows:
            t = task_map.get(row["task_id"])
            if t:
                row["pass1_prompt"] = t.pass1_prompt or ""
                row["transcript"] = t.raw_transcript or ""

        wb = openpyxl.Workbook()

        # ── Sheet 1: Results ──────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Results"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F46E5")  # indigo (used in Personas sheet)

        if rows:
            headers = list(rows[0].keys())

            # Colour scheme per column group:
            #   indigo  = metadata (task/persona/status/cost)
            #   teal    = injected variables (var_*)
            #   green   = extracted outputs
            #   purple  = confidence scores (*_confidence)
            #   grey    = transcript
            meta_fill    = PatternFill("solid", fgColor="4F46E5")   # indigo
            var_fill     = PatternFill("solid", fgColor="0F766E")   # teal
            output_fill  = PatternFill("solid", fgColor="15803D")   # green
            conf_fill    = PatternFill("solid", fgColor="7E22CE")   # purple
            other_fill   = PatternFill("solid", fgColor="6B7280")   # grey

            meta_cols = {"task_id", "persona_id", "pass1_status", "pass2_status",
                         "drift_flagged", "pass1_cost_usd", "pass2_cost_usd", "calibration_badge"}

            for col_idx, h in enumerate(headers, 1):
                if h in meta_cols:
                    fill = meta_fill
                elif h.startswith("var_"):
                    fill = var_fill
                elif h.endswith("_confidence"):
                    fill = conf_fill
                elif h in ("transcript", "pass1_prompt"):
                    fill = other_fill
                else:
                    fill = output_fill

                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center")

            for row in rows:
                ws.append([row.get(h) for h in headers])

            # Auto-width columns (cap at 80 for transcript)
            for col_idx, h in enumerate(headers, 1):
                col_letter = get_column_letter(col_idx)
                max_len = max(len(str(h)), max(
                    len(str(r.get(h, "") or "")) for r in rows
                ))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

            # Wrap text for long text columns (pass1_prompt, transcript)
            for long_col_name in ("pass1_prompt", "transcript"):
                if long_col_name not in headers:
                    continue
                col_idx_long = headers.index(long_col_name) + 1
                col_letter = get_column_letter(col_idx_long)
                ws.column_dimensions[col_letter].width = 60
                for row_idx in range(2, len(rows) + 2):
                    ws.cell(row=row_idx, column=col_idx_long).alignment = Alignment(
                        wrap_text=True, vertical="top"
                    )

        ws.freeze_panes = "A2"

        # ── Sheet 2: Personas ─────────────────────────────────────────────────
        persona_ids = list({task.persona_id for task in tasks})
        from ..models.audience import Persona as PersonaModel
        from sqlalchemy import select as sa_select
        personas_result = await db.execute(
            sa_select(PersonaModel).where(PersonaModel.id.in_(persona_ids))
        )
        personas = {p.id: p for p in personas_result.scalars().all()}

        ws_p = wb.create_sheet(title="Personas")

        # Collect all trait keys across all personas
        trait_keys: list[str] = []
        for pid in persona_ids:
            p = personas.get(pid)
            if p and p.traits_json:
                for k in p.traits_json:
                    if k not in trait_keys:
                        trait_keys.append(k)

        persona_headers = ["persona_id", "audience_id", "plausibility", "flagged", "backstory"] + trait_keys
        for col_idx, h in enumerate(persona_headers, 1):
            cell = ws_p.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for pid in persona_ids:
            p = personas.get(pid)
            if not p:
                continue
            traits = p.traits_json or {}
            row_vals = [
                p.id,
                p.audience_id,
                p.plausibility,
                p.flagged,
                p.backstory or "",
            ] + [traits.get(k, "") for k in trait_keys]
            ws_p.append(row_vals)

        # Auto-width persona columns
        for col_idx, h in enumerate(persona_headers, 1):
            col_letter = get_column_letter(col_idx)
            if h == "backstory":
                ws_p.column_dimensions[col_letter].width = 60
                for row_idx in range(2, len(persona_ids) + 2):
                    ws_p.cell(row=row_idx, column=col_idx).alignment = Alignment(
                        wrap_text=True, vertical="top"
                    )
            else:
                max_len = max(len(str(h)), max(
                    (len(str(personas[pid].traits_json.get(h, "") if h in trait_keys
                          else getattr(personas[pid], h, "") or ""))
                     for pid in persona_ids if pid in personas),
                    default=10,
                ))
                ws_p.column_dimensions[col_letter].width = min(max_len + 2, 40)

        ws_p.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=run_{run_id[:8]}.xlsx"},
        )
